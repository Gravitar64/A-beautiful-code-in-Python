import glob
from openpyxl import Workbook, load_workbook
import re

pfad = "Teil_059_excel/"
dateien = glob.glob(pfad + "Prio_*.xlsx")
z_wb = Workbook()
z_ws = z_wb.active

sp_themen = "A"
sp_punkte = "B"

thema2zeile = {}
for n, datei in enumerate(dateien):
  teiln = re.findall("_[\w]+", datei)[1][1:]
  q_wb = load_workbook(datei)
  q_ws = q_wb.active
  themen = [x.value for x in q_ws[sp_themen]]
  punkte = [x.value for x in q_ws[sp_punkte]]
  p_int = [x for x in punkte if type(x) == int and 0 < x < 4]
  if sum(p_int) != 6 or not {1, 2, 3}.issubset(p_int):
    print(f'Rückmeldung von {teiln} ist fehlerhaft {punkte}')
    continue
  if not thema2zeile:
    thema2zeile = {thema: zeile + 1 for zeile, thema in enumerate(themen)}
    for thema in zip(themen):
      z_ws.append(thema)
  for thema, punkt in zip(themen, punkte):
    z_ws.cell(thema2zeile[thema], n + 2).value = punkt
  z_ws.cell(1, n + 2).value = teiln

z_wb.save(pfad + "Ergebnis_gesamt.xlsx")
