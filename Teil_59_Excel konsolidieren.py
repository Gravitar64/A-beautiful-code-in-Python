import glob, os
import re
from openpyxl import load_workbook, Workbook


dateien = glob.glob("Teil_59_excel/*.xlsx")
z_wb = Workbook()
z_ws = z_wb.active
sp_themen = "A"
sp_punkte = "B"

p2t = {}
for n,datei in enumerate(dateien):
  teiln = re.findall("_[\w]+", datei)[1][1:]
  q_wb = load_workbook(datei)
  q_ws = q_wb.active
  t = [x.value for x in q_ws[sp_themen]]
  p = [x.value for x in q_ws[sp_punkte]]
  p_int = [x for x in p if type(x)==int and 0<x<4]
  if sum(p_int) != 6 or not {1,2,3}.issubset(p_int):
    print(f'Rückmeldung von {teiln} nicht korrekt {p}')
    continue
  if not p2t:
    p2t = {thema:n+1 for n,thema in enumerate(t)}
    for thema in zip(t):
      z_ws.append(thema)
  for thema,punkt in zip(t,p):
    if type(punkt) != int or punkt > 3: continue
    z_ws.cell(p2t[thema],n+2).value = punkt
    z_ws.cell(1,n+2).value = teiln

z_wb.save("Teil_59_ergebnis.xlsx")    

